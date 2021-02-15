# Import module
from tkinter import *
import requests
import time
import tkinter
from bs4 import BeautifulSoup
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
import pandas as pd
import requests
import json
import threading
import re
import csv
from tldextract import extract
from openpyxl import load_workbook
import os
from urllib.request import urlopen
import tkinter.messagebox
from tkinter import Tk, Text, Scrollbar
from tkinter import Tk, Label, OptionMenu, StringVar, Button, font




#https://gothiclandscape.com/
#https://www.marianilandscape.com/where-we-are/
#https://www.yellowstonelandscape.com/locations -NA
# https://www.ruppertlandscape.com/     -----Forbidden
# https://www.schilllandscaping.com/   no locations
#https://www.junipercares.com/
# https://sebert.com/       no locations
# https://uslawns.com/
#https://landscapedevelopment.com/contact/
def get_options(scrambled, flag, totals, last):
    dictionary = [i.strip('\n') for i in open('the_file.txt')]
    if flag:
        return totals
    else:
        new_list = [i for i in dictionary if scrambled.startswith(i)]
        if new_list:
            possible_word = new_list[-1]
            new_totals = totals
            new_totals.append(possible_word)
            new_scrambled = scrambled[len(possible_word):]
            return get_options(new_scrambled, False, new_totals, possible_word)
        else:
            return get_options("", True, totals, '')
class locations:
    #https://www.marianilandscape.com/where-we-are/
    def child_tree1(url, soup, td, branches):
       
        tsd, td, tsu = extract(url)
        to_csv = []
        pattern = re.compile(':(\[.*?\}\])')
        try:
            for script in soup.find_all('script',  {"type": "text/javascript"}):
                if len(script.contents) > 0:
                    if 'address' in script.contents[0]:
                        match = pattern.search(script.string)
                        if match is not None:
                            result = json.loads(match.groups()[0])
            for d in result:
                sample_text = d['address']
                us_zip = r'(\d{5}\-?\d{0,4})'
                zip_code = re.search(us_zip, sample_text)
                zip_code = zip_code.group(1)
                sample_text = sample_text.replace(zip_code, "")
                sample_text = sample_text.replace(", ,", ",")
                if sample_text[-2].endswith(','):
                    sample_text = sample_text[0:len(sample_text)-2]
                # extracting entities.
                csv_list = [td, d['title'], branches, sample_text, sample_text.split(',')[-2],
                                sample_text.split(',')[-1] , zip_code, d['email'],  url]
                to_csv.append(csv_list)
            return to_csv
        except:
             return None
    #https://landscapedevelopment.com/contact/
    def child_tree2(url, soup, td, branches):
        tsd, td, tsu = extract(url)
        to_csv = []
        try:
            for script in soup.find_all('script',  {"type": "application/ld+json"}):
                try:
                    address = json.loads(script.string)
                    result = address['address']
                except:
                    pass
            for d in result:
                csv_list = [td.capitalize(), d['name'].capitalize(), branches, d['streetAddress'], d['addressLocality'],
                                        d['addressRegion'], d['postalCode'] , d['telephone'],  d["hasMap"]]
                to_csv.append(csv_list)
            return to_csv
        except:
            return None
   # https://www.yellowstonelandscape.com/locations
    def child_tree3(url, soup, td, branches):
        tsd, td, tsu = extract(url)
        to_csv = []
        pattern = re.compile('(\[\{.*?\}\])')  
        try:    
            for script in soup.find_all("script", {"src":False}):
                if script:
                    m = pattern.search(script.string)
                    if m is not None:
                        result = json.loads(m.groups()[0])
               
            for d in result:
                csv_list = [td, "NA", branches, d['address'], d['city'],d['state'], d['zip'], d['phone'],  d['link']]
                to_csv.append(csv_list)
            return to_csv
        except:
            return None
     #https://www.junipercares.com/
    def child_tree4(url, soup, td, branches):
        tsd, td, tsu = extract(url)
        to_csv = []
        pattern = re.compile('(\'.*?\')')
        if len(soup.find_all('script',  {"type": "text/javascript"})) == 0:
            return None
        try:
            for script in soup.find_all('script',  {"type": "text/javascript"}):
                if script.string is None:
                    continue
                try:
                    match = pattern.search(script.string)
                    var = match.groups()[0]
                    result = var.replace('\\u003Cstrong\\u003E','').replace('\\u0020', ' ').replace('\\u003C\\/strong\\u003E\\u003Cbr\\u003E', ",").replace('u003Cbr\\u003E',",").replace('\\nPh\\u003A \\u0028', '').replace('\\u0029', '-').replace('\\', '').replace("u002D3", '-').replace('u002D5',  '').replace('u002D7' , '').replace('u002D6',  '')      
                except:
                    continue
                if len(result) > 20:
                    result =  result.replace(', ,', ',')
                    result = result.split(',')
                    us_zip = r'(\d{5}\-?\d{0,4})'
                    zip_code = re.search(us_zip, result[3].strip())
                    region = result[3].replace(zip_code.group(1), '')
                    sub_branch = url.split('/')
                    csv_list = [td, sub_branch[-1].replace('-', ' ').capitalize(), branches, result[1], result[2],
                    region.strip(), zip_code.group(1), result[-1].replace("'", ''),  url]
                    to_csv.append(csv_list)
                    if len(to_csv) != 0:
                        return to_csv
           
                    break
               
        except:
            return None
    #https://gothiclandscape.com/
    def child_tree5(url, soup, td, branches):
        to_csv = []
        link_site = []
        r = requests.get(url, verify = False)
        soup = BeautifulSoup(r.content, 'html.parser')
        for link in soup.find_all('a'):
            link  = link.get('href')
            if link is None:
                continue
            if 'branch' in link or 'location/' in link:
                if 'https' not in link:
                    link = url + link
                link_site.append(link)
        link_site = list(set(link_site))
        for url in link_site:
            r = requests.get(url, verify = False)
            soup = BeautifulSoup(r.content, 'html.parser')
            if 'california' in url:
                 place = ['col-24 c-py-9 border-top border-grey', 'col-24 c-py-9']
            else:
                place = ["col-lg-12 c-py-9", "col-lg-12 c-py-9 border-top border-grey border-lg-none"]
            for i in place:
                all_div = soup.find_all("div", class_ = i)
                for div in all_div:
                    data = div.text.split('\n')
                result = []
                datastore = {}
                for i in data:
                    if 'Landscape' in i:
                        sub_branch = i
                        continue
                    if len(i) > 0:
                        if '.com' not in str(i):
                            datastore[i] = i
                        else:
                            datastore[i] = i
                            result.append(datastore)
                            datastore = {}
                for d in result:
                    if 'arizona' in url or 'nevada' in url:
                        address = list(d.keys())[2]
                        address= address.split(',')
                        us_zip = r'(\d{5}\-?\d{0,4})'
                        zip_code = re.search(us_zip, address[-1])
                        phone = list(d.keys())[3].replace('Phone: ', '')
                        state = address[-1].replace(zip_code.group(1), '').strip()
                        csv_list = [td.capitalize(), sub_branch.capitalize(), branches.capitalize(), address[0], list(d.keys())[1],
                                                        state, zip_code.group(1), phone.replace('Management Contact:', ''),  url]
                        to_csv.append(csv_list)
                    else:
                        address=list(d.keys())[3]
                        address= address.split(',')
                        us_zip = r'(\d{5}\-?\d{0,4})'
                        zip_code = re.search(us_zip, address[-1])
                        phone = list(d.keys())[4].replace('Phone: ', '').strip()
                        phone_tt = phone.replace('-', '')
                        if not phone_tt.isdigit():
                            phone = 'NA'
                        state = address[-1].replace(zip_code.group(1), '').strip()
                        csv_list = [td.capitalize(), sub_branch.capitalize(), branches.capitalize(), address[0], list(d.keys())[2],
                                                        state, zip_code.group(1), phone.replace('Management Contact:', ''),  url]
                        to_csv.append(csv_list)
                    if len(d) > 10:
                        if 'arizona' in url or 'nevada' in url:
                            address=list(d.keys())[5]
                            address= address.split(',')
                            us_zip = r'(\d{5}\-?\d{0,4})'
                            zip_code = re.search(us_zip, address[-1])
                            phone = list(d.keys())[6].replace('Phone: ', '')
                            state = address[-1].replace(zip_code.group(1), '').strip()
                            csv_list = [td.capitalize(), sub_branch.capitalize(), branches.capitalize(), address[0], list(d.keys())[4],
                                                        state, zip_code.group(1), phone,  url]
                            to_csv.append(csv_list)
                        else:
                           
                            address=list(d.keys())[6]
                            address= address.split(',')
                            us_zip = r'(\d{5}\-?\d{0,4})'
                            zip_code = re.search(us_zip, address[-1])
                            phone = list(d.keys())[8].replace('Phone: ', '').strip()
                            phone_tt = phone.replace('-', '')
                            if not phone_tt.isdigit():
                                phone = 'NA'
                            state = address[-1].replace(zip_code.group(1), '').strip()
                            csv_list = [td.capitalize(), sub_branch.capitalize(), branches.capitalize(), address[0], list(d.keys())[5],
                                                        state, zip_code.group(1), phone,  url]
       
                            to_csv.append(csv_list)
        if len(to_csv) > 0:
            return to_csv
        else:
            return None
   
   
    #https://www.ruppertlandscape.com/
    def child_tree6(url, soup):
        print('inside child_tree 6')
        to_csv = []
        service=[]
        h1 ='h1'
        h2 = 'h2'
        h3 ='h3'
        h4 ='h4'
        adder = []
        tsd, td, tsu = extract(url)
        td = td.capitalize()
        heading_tags = ["h1", "h2", "h3" , "h4"]
        for tags in soup.find_all(heading_tags):
            #print(tags.name + ' -> ' + tags.text.strip() +  ' -> ')
            service.append(tags.text.strip())
       
        class_list = set()
        # get all tags
        tags = {tag.name for tag in soup.find_all()}
        # iterate all tags
        for tag in tags:
            # find all element of tag
            for i in soup.find_all( tag ):
                # if tag has attribute of class
                if i.has_attr( "class" ):
                    if len( i['class'] ) != 0:
                        class_list.add(" ".join( i['class']))
        #print(class_list)
        for each_class in class_list:
            all_division  = soup.find_all("div",class_= each_class)
            tag = list(set([str(tag.name) for tag in soup.find_all()]))
            all_tags = []
            for i in tag:
                if ('p' in i or 'h' in i) and len(i) <= 2:
                    all_tags.append(i)
            heading_tags = ["h1", "h2", "h3", "h4"]      
            for all_div in all_division:
                if h1 in all_tags:
                    try:
                        if all_div.h1.text in service and all_div.h1.text not in adder:
                            #adder.append(all_div.h1.text)
                            #print(list(zip([all_div.h1.text], [all_div.p.text])))
                            #df = pd.DataFrame(list(zip([all_div.h1.text], [all_div.p.text])) ,columns = cols)
                            sample_text = all_div.p.text
                            address = sample_text.split('\n')[0].split(',')[0]
                            textlabel = sample_text.split('\n')[0].split(',')[1].split('P:')[0]
                            phone= sample_text.split('\n')[0].split(',')[1].split('P:')[-1]
                            state = ''
                            zipcode = ''                            
                            for i in textlabel:
                                if i.isdigit():
                                    zipcode +=i
                                else:
                                    state +=i
                            csv_list = [td, "NA", branches, address, "",state, zipcode, phone,  url]
                            to_csv.append(csv_list)
                    except:
                        pass
                if h2 in all_tags:
                    try:
                        if all_div.h2.text in service and all_div.h2.text not in adder:
                            #adder.append(all_div.h2.text)
                            #print(list(zip([all_div.h2.text], [all_div.p.text])))
                            #df = pd.DataFrame(list(zip([all_div.h2.text], [all_div.p.text])) ,columns = cols)
                            sample_text = all_div.p.text
                            address = sample_text.split('\n')[0].split(',')[0]
                            textlabel = sample_text.split('\n')[0].split(',')[1].split('P:')[0]
                            phone= sample_text.split('\n')[0].split(',')[1].split('P:')[-1]
                            state = ''
                            zipcode = ''                            
                            for i in textlabel:
                                if i.isdigit():
                                    zipcode +=i
                                else:
                                    state +=i
                            csv_list = [td, "NA", branches, address, "",state, zipcode, phone,  url]
                            to_csv.append(csv_list)
                    except:
                        pass
                if h3 in all_tags:
                    try:
                        if all_div.h3.text in service and all_div.h3.text not in adder:
                            #adder.append(all_div.h3.text)
                            #print(list(zip([all_div.h3.text], [all_div.p.text])))
                            #df = pd.DataFrame(list(zip([all_div.h3.text], [all_div.p.text])) ,columns = cols)
                            sample_text = all_div.p.text
                            address = sample_text.split('\n')[0].split(',')[0]
                            textlabel = sample_text.split('\n')[0].split(',')[1].split('P:')[0]
                            phone= sample_text.split('\n')[0].split(',')[1].split('P:')[-1]
                            state = ''
                            zipcode = ''                            
                            for i in textlabel:
                                if i.isdigit():
                                    zipcode +=i
                                else:
                                    state +=i
                            csv_list = [td, "NA", branches, address, "",state, zipcode, phone,  url]
                            to_csv.append(csv_list)
                    except:
                        pass
                if h4 in all_tags:
                    try:
                        if all_div.h4.text in service and all_div.h4.text not in adder:
                            #adder.append(all_div.h4.text)
                            #print(list(zip([all_div.h4.text], [all_div.p.text])))
                            #df = pd.DataFrame(list(zip([all_div.h4.text], [all_div.p.text])) ,columns = cols)
                            sample_text = all_div.p.text
                            address = sample_text.split('\n')[0].split(',')[0]
                            textlabel = sample_text.split('\n')[0].split(',')[1].split('P:')[0]
                            phone= sample_text.split('\n')[0].split(',')[1].split('P:')[-1]
                            state = ''
                            zipcode = ''                            
                            for i in textlabel:
                                if i.isdigit():
                                    zipcode +=i
                                else:
                                    state +=i
                            csv_list = [td, "NA", branches, address, "",state, zipcode, phone,  url]
                            to_csv.append(csv_list)
                   
                    except:
                        pass
       
       
        if len(to_csv) > 0:
            return to_csv
        else:
            return None

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
       # Excel file doesn't exist - saving and exiting
    if not filename.endswith('xlsx'):
        filename = filename + '.xlsx'
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return
   
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
   
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly.
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
   
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()    
   
def location_parser(urls):
    to_csv = []
    link_site = []
    for given_url in [urls]:
        url = given_url
        if not given_url.endswith('locations') and given_url.endswith('com/'):
            if given_url[-1] != '/':
                given_url = given_url + '/locations'
            else:
                given_url = given_url + 'locations'
        print('inside location_parser', urls)
        r = requests.get(given_url, verify = False)
        if r.status_code != 200 and not 'contact' in url :
            r = requests.get(url, headers={"Content-Type":"application/json",
                                        "User-Agent":"PostmanRuntime/7.29.2",
                                        "Accept":"*/*"}, verify = False)
            soup = BeautifulSoup(r.content, 'html.parser')
            for link in soup.find_all('a'):
                link  = link.get('href')
                if link is None or 'selected' in link:
                    continue
                if 'branch' in link or 'locations' in link:
                    if not link.startswith('https:'):
                        link = url + link
                        link_site.append(link)
                    else:
                        link_site.append(link)
        soup = BeautifulSoup(r.content, 'html.parser')
        tsd, td, tsu = extract(url)
        td  = ' '.join(get_options(td, False, [], '')).capitalize()
        branch_type = ["maintenance", "installation", "enhancements", "irrigation", "snow-ice",
                       "design", "construct", "commercial", "maintain", "landscape-construction", "landscape-management"
                  "environmental-restoration", "sustainable-environments"]
        branch_type_string = "".join(branch_type)
        branches = ""
        for link in soup.find_all('a'):
            link  = link.get('href')
            if link is None:
                continue
            each_href = link.split('/')
            for href in each_href:
                if any(ext in href for ext in branch_type):
                    before, sep, after = href.partition('?')
                    before  = re.sub('[^A-Za-z0-9]+', ' ', before)
                   # before  = ' '.join(get_options(before, False, [], ''))
                    before = before.capitalize()
                    if 'jpg' in before:
                        continue
                    if 'com' in before:
                        if 'comm' in before:
                            pass
                        else:
                            continue
                    branches= branches + before + ","
        final_branch = set(branches.split(','))
        final_branch = ",".join(final_branch)
        branches = final_branch.replace(',,', ',')
        link_site =list(set(link_site))
        multisite = []
        if len(link_site) > 1:
            for url in link_site:
                r = requests.get(url, verify = False)
                soup = BeautifulSoup(r.content, 'html.parser')
                methods = [getattr(locations, m) for m in dir(locations) if not m.startswith('__')]
                for method in methods:
                    try:
                      data = method(url, soup, td, branches)
                    except:
                        # Can't handle methods with required arguments.
                        pass
                    if data is None:
                        pass
                    else:
                        multisite.append(data[0])
            to_csv.append(multisite)
        else:
            methods = [getattr(locations, m) for m in dir(locations) if not m.startswith('__')]
            for method in methods:
                try:
                    if len(link_site) > 0:
                        r = requests.get(link_site[0], headers={"Content-Type":"application/json",
                                            "User-Agent":"PostmanRuntime/7.29.2",
                                            "Accept":"*/*"}, verify = False)
                        soup = BeautifulSoup(r.content, 'html.parser')
                        data = method(link_site[0], soup)
                    else:
                        data = method(url, soup, td, branches)
                except TypeError:
                    # Can't handle methods with required arguments.
                    pass
                if data is None:
                    pass
                else:
                    to_csv.append(data)
    return to_csv                
                   

def single_url(soup):
    adder = []
    heading_tags = ["h1", "h2", "h3" , "h4"]
    for tags in soup.find_all(heading_tags):
        #print(tags.name + ' -> ' + tags.text.strip() +  ' -> ')
        service.append(tags.text.strip())

    class_list = set()
    # get all tags
    tags = {tag.name for tag in soup.find_all()}
    # iterate all tags
    for tag in tags:
        # find all element of tag
        for i in soup.find_all( tag ):
            # if tag has attribute of class
            if i.has_attr( "class" ):
                if len( i['class'] ) != 0:
                    class_list.add(" ".join( i['class']))
    #print(class_list)
    for each_class in class_list:
        all_division  = soup.find_all("div",class_= each_class)
        tag = list(set([str(tag.name) for tag in soup.find_all()]))
        all_tags = []
        for i in tag:
            if ('p' in i or 'h' in i) and len(i) <= 2:
                all_tags.append(i)
        heading_tags = ["h1", "h2", "h3"]      
        for all_div in all_division:
            if h1 in all_tags:
                try:
                    if all_div.h1.text in service and all_div.h1.text not in adder:
                        adder.append(all_div.h1.text)
                        #print(list(zip([all_div.h1.text], [all_div.p.text])))
                        df = pd.DataFrame(list(zip([all_div.h1.text], [all_div.p.text])) ,columns = cols)
                        append_df_to_excel(output, df, all_div.h1.text, index=False)
                except:
                    pass
            if h2 in all_tags:
                try:
                    if all_div.h2.text in service and all_div.h2.text not in adder:
                        adder.append(all_div.h2.text)
                        #print(list(zip([all_div.h2.text], [all_div.p.text])))
                        df = pd.DataFrame(list(zip([all_div.h2.text], [all_div.p.text])) ,columns = cols)
                        append_df_to_excel(output, df, all_div.h2.text, index=False)
                except:
                    pass
            if h3 in all_tags:
                try:
                    if all_div.h3.text in service and all_div.h3.text not in adder:
                        adder.append(all_div.h3.text)
                        #print(list(zip([all_div.h3.text], [all_div.p.text])))
                        df = pd.DataFrame(list(zip([all_div.h3.text], [all_div.p.text])) ,columns = cols)
                        append_df_to_excel(output, df, all_div.h3.text, index=False)
                except:
                    pass

# Change the label text.
def get_categories():
    if entry1.get():
        url = entry1.get()
        categories = ["about","privacy-policy","jobs", "portfolio", "customers","where-we-are","locations",
                     "services", "portfolio", "careers", "projects", "branch",
                    "branches", "locations", "who-we-serve", "contact-us",
                  "news", "acquisitions", "testimonials", "case-studies","resources", "about-us"]
        c = []
        if url.startswith('http') or url.startswith('wwww'):
            response = requests.get(url,
                headers={"Content-Type":"application/json",
                "User-Agent":"PostmanRuntime/7.29.2",
                "Accept":"*/*"},
                verify = False)
            soup = BeautifulSoup(response.content, 'html.parser')
            for link in soup.find_all('a'):
                link  = link.get('href')
                if link is None:
                    continue
                for i in categories:
                    if i in link and i not in c:
                        c.append(i)
        c = list(set(c))
        #helv36 = tkFont.Font(family='arial', size=10)
        main_menu = tkinter.OptionMenu(root, clicked, 'home', *c)
        #main_menu.config(font=helv36)
        main_menu.place(x= 120, y =260)
        main_menu.config(width = 15)
        #main_menu.pack()
        return c

def execute_script():
    def sub_execute():
        print("Wait 5 seconds")
        time.sleep(5)
        if entry1.get():
           new_urls = []
           dropdown_value = clicked.get()
           given_url = entry1.get()
           output = entry2.get()
           print('dropdown')
           if dropdown_value == 'locations' or 'branch' in dropdown_value:
               #myscript = threading.Thread(target=location_parser, args=(given_url,))
               #print(given_url)
               #myscript.start()
               list_data = location_parser(given_url)
               fields =  ["Parent Company", "Sub-Brand (if relevant)",
                          "Branch Type (If relevant)", "Street", "City", "State", "Zip Code", "Other Comments", "Link"]
               df =pd.DataFrame(list_data[0], columns = fields)
               time.sleep(5)
               append_df_to_excel(output, df, 'Locations', index = False)
               tkinter.messagebox.showinfo("Web scrapping result", "Successfully processed")    
                         
           if dropdown_value == "Yelp reviews":
               print('inside Yelp')
               if given_url.startswith('http') or given_url.startswith('www'):
                   tsd, td, tsu = extract(given_url)
                   td  = ' '.join(get_options(td, False, [], '')).capitalize()
               else:
                   td = given_url
               
               td = td.replace(' ', "%20")
               url = "https://maps.googleapis.com/maps/api/place/findplacefromtext/json?input="
               url3= "&inputtype=textquery&fields=place_id&key=AIzaSyDMg-POYO8aPVIX4rOhUZqGus4GfdoMtXE"
               url = url+td+url3
               payload={}
               headers = {}
               response = requests.request("GET", url, headers=headers, data=payload)
               response = response.json()
               try:
                   response = response['candidates'][0]['place_id']
                   init_url= "https://maps.googleapis.com/maps/api/place/details/json?placeid="
                   api_key = "&key=AIzaSyDMg-POYO8aPVIX4rOhUZqGus4GfdoMtXE"
                   review_url = init_url + response + api_key
                   review_response = requests.request("GET", review_url, headers=headers, data=payload)
                   res = review_response.json()['result']['adr_address']
                   html = res.replace(res, res)
                   soup = BeautifulSoup(html, "html.parser")
                   print('beautiful soup')
                   address1 = soup.find("span", class_="street-address").text.replace(' ', '%20')
                   address1 = address1.replace('#', "%23")
                   city  = soup.find("span", class_="locality").text.replace(' ', '%20')
                   state = soup.find("span", class_="region").text
                   country = soup.find("span", class_="country-name").text[0:2]
                   yelp_url = "https://api.yelp.com/v3/businesses/matches?"
                   name = "name={0}".format(td)
                   address1 =  "&address1={0}".format(address1)
                   city  = "&city={0}".format(city)
                   state  = "&state={0}".format(state)
                   country = "&country={0}".format(country)
                   yelp_api_key= "IEDZzra9jWr5BBJhZJkUV2_ck3ozOFjL8sIuCakAPWBqN6gKPN8R2wkliVO97WsCWp7QRos_P9m_ToRtTn4tpPaN77HkCyn1huepegGpVlqUNrAZCqwi5lKj8xw0Y3Yx"
                   yelp_url = yelp_url + name + address1 + city + state + country +"&limit=3&match_threshold=default"
                   print(yelp_url)
                   yelp_headers = {'Authorization': 'Bearer %s' % yelp_api_key,}
                   yelp_response = requests.request("GET",yelp_url, headers=yelp_headers, verify=False)
                   place_id = yelp_response.json()['businesses'][0]['id']
                   url = "https://api.yelp.com/v3/businesses/"
                   part_url = place_id + "/reviews?limit=20&sort_by=yelp_sort"
                   url = url + part_url
                   response = requests.request('GET', url, headers=yelp_headers, verify=False)
                   #response = requests.request('GET', url, headers=headers, params=url_params, verify=False)
                   print(response.status_code)
                   store_data = []
                   response = response.json()['reviews']
                   print(response)
                   for i in response:
                       processed_item = {}
                       processed_item['Username'] = i['user']['name']
                       processed_item['URL']  = i['url']
                       processed_item['Comments']   = i['text']
                       processed_item['Rating']  = i['rating']
                       processed_item['Created Time'] = i['time_created']
                       store_data.append(processed_item)

                   df = pd.DataFrame(store_data)
                   cols = list(df.columns)
                   new_cols = []
                   for col in cols:
                       new_cols.append(col.capitalize())
                   df.columns = new_cols
                   output = str(output)
                   if output.endswith('xlsx'):
                       pass
                   else:
                       output =output +".xlsx"
                   append_df_to_excel(output, df, 'Reviews', index = False)
                   tkinter.messagebox.showinfo("Web scrapping result", "Successfully processed")  
               
               except:
                   tkinter.messagebox.showinfo("Web scrapping result", "Reviews not found")

               
           
           else:
               tsd, td, tsu = extract(given_url)
               response = requests.get("https://www." + td +".com/",
                                        headers={"Content-Type":"application/json",
                                                "User-Agent":"PostmanRuntime/7.29.2",
                                                "Accept":"*/*"},
                                        verify = False)
               soup = BeautifulSoup(response.content, 'html.parser')
               print('before for loop')
               for link in soup.find_all('a'):
                   link  = link.get('href')
                   if link is None or 'unhappy' in link:
                        continue        
                   if not link.startswith('http'):
                       if link.startswith('/'):
                           url = "https://" + td + ".com"
                       else:
                           url = "https://" + td + ".com"
                       link  = url + link
                   if ('services' in link or 'service' in link or 'our-work' in link) and len(link.split('/'))>=5:
                       if 'http' in link:
                           link = link.replace("?hsLang=en", "")
                           link = link.replace("sebertwi", "sebert")
                           try:
                               result = urlopen(link)
                               link = result.geturl()
                           except:
                               pass
                           new_urls.append(link)
       
               new_urls = list(set(new_urls))
               if len(new_urls) == 0:
                   if given_url[-1] != '/':
                       url = given_url + '/services'
                   else:
                       url = given_url + 'services'
                   response = requests.get(url, headers={"Content-Type":"application/json",
                                        "User-Agent":"PostmanRuntime/7.29.2",
                                        "Accept":"*/*"}, verify = False)
                   if response.status_code != 200:
                       commercial_url = given_url + "commercial/"
                       response = requests.get(given_url, headers={"Content-Type":"application/json",
                                        "User-Agent":"PostmanRuntime/7.29.2",
                                        "Accept":"*/*"}, verify = False)
                       if response.status_code != 200:
                           landscape_url = given_url + "landscaping-services//"
                           response = requests.get(given_url, headers={"Content-Type":"application/json",
                                        "User-Agent":"PostmanRuntime/7.29.2",
                                        "Accept":"*/*"}, verify = False)  
                   soup = BeautifulSoup(response.content, 'html.parser')
                   single_url(soup)
               multiurl =[]
               if len(new_urls) > 3:
                   for i in new_urls:
                       if i.endswith('services/') or i.endswith('service/') or i.endswith('services') or i.endswith('service/') or 'hire' in i:
                           pass
                       else:
                           multiurl.append(i)
                   list_of_dataframes  = []
                   for sub_url in multiurl:
                       response = requests.get(sub_url, headers={"Content-Type":"application/json",
                                            "User-Agent":"PostmanRuntime/7.29.2",
                                            "Accept":"*/*"}, verify = False)
                       url_split = sub_url.split('/')
                       if url_split[-1] == "":
                           url_element = url_split[-2]
                       else:    
                           url_element = url_split[-1]
                       url_element = url_element.replace('-', ' ')
                       sheet_name = url_element.capitalize()
                       soup = BeautifulSoup(response.content, 'html.parser')
                       heading_tags = ["h1", "h2", "h3" , "h4"]
                       class_list = set()
                       # get all tags
                       tags = {tag.name for tag in soup.find_all()}
                       # iterate all tags
                       for tag in tags:
                            # find all element of tag
                           for i in soup.find_all(tag):
                                # if tag has attribute of class
                                if i.has_attr("class"):
                                    if len( i['class'] ) != 0:
                                        class_list.add(" ".join( i['class']))
                       description = []
                       data_creation = []
                       sheet_creation = []
                       for each_class in class_list:
                           all_division  = soup.find_all("div", class_= each_class)
                           for all_div in all_division:
                               try:
                                    page = all_div.find_all('p')
                                    text = [p.get_text() for p in page]
                                    for i in text:
                                        if i not in description:
                                            description.append(i)
                               except:
                                   pass
                       
                       description = "".join(description)
                       description = description.lstrip()
                       data_creation.append(description)
                       sheet_creation.append(sheet_name)
                       cols = ['Services', 'Description']
                       df = pd.DataFrame(list(zip(sheet_creation, data_creation)) ,columns = cols)
                       list_of_dataframes.append(df)
                       #print(list_of_dataframes)
                   myDataFrame  = pd.concat(list_of_dataframes, axis=0, ignore_index=True)
                   append_df_to_excel(output, myDataFrame, 'Services', index = False)          
                   
                   tkinter.messagebox.showinfo("Web scrapping result", "Successfully processed")    
               
               else:
                   response = requests.get(new_urls[0], headers={"Content-Type":"application/json",
                                        "User-Agent":"PostmanRuntime/7.29.2",
                                        "Accept":"*/*"}, verify = False)
                   soup = BeautifulSoup(response.content, 'html.parser')
                   single_url(soup)
               
    threading.Thread(target=sub_execute).start()
    #root.mainloop()
   
def openstreet_execute_script():
    try:
        results =  []
        url = entry1.get()
        print(url)
        response = requests.get(url, headers={"Content-Type":"application/json",
                                            "User-Agent":"PostmanRuntime/7.29.2",
                                            "Accept":"*/*"}, verify = False)
        soup = BeautifulSoup(response.content, 'html.parser')
        heading_tags = ["h1", "h2", "h3" , "h4"]
        class_list = set()
        # get all tags
        tags = {tag.name for tag in soup.find_all()}
        # iterate all tags
        print('before for')
        for tag in tags:
            for i in soup.find_all(tag):
                if i.has_attr("class"):
                    if len( i['class'] ) != 0:
                        class_list.add(" ".join( i['class']))
        #print(class_list)
        all_division  = soup.find_all("span", class_= "latitude")
        for all_div in all_division:
            lat = all_div.text
        all_division  = soup.find_all("span", class_= "longitude")
        for all_div in all_division:
            long = all_div.text
        all_division  = soup.find_all("table", class_= "browse-tag-list border border-grey rounded")
        for all_div in all_division:
            try:
                value = str(all_div.text).split('\n')
                value = [i for i in value if i!='']
                keys =  []
                values = []
                for i in range(len(value)):
                    num = i+1
                    if (num % 2) == 0:
                        values.append(value[i].capitalize())
                    else:
                        keys.append(value[i].capitalize())
                res = dict(zip(keys, values))    
            except:
                pass

        print(res)
        df = pd.DataFrame([res])
        print(df.head())
        df = df.fillna('')
        output = entry2.get()
        if output.endswith('xlsx'):
            pass
        else:
            output =output +".xlsx"
        append_df_to_excel(output, df, 'Map_data', index = False)
        tkinter.messagebox.showinfo("Web scrapping result", "Successfully processed")  
             
    except:
        tkinter.messagebox.showinfo("Web scrapping result", "Error occured")
   
   
def open_street_reviews():
    global entry1, entry2
    label1 = Label(root,text="Enter the map URL",font=('arial',16),fg="black")
    label1.place(x=150, y=140)
    entry1 = Entry(root, width=40)
    entry1.place(x=150, y=180)
   
    label2 = Label(root,text="Enter name to save file",font=('arial',16),fg="black")
    label2.place(x=150,y=210)
    entry2 = Entry(root, width=40)
    entry2.place(x=150, y=240)
   
    button = Button(root,width=10, text=" Download",command = openstreet_execute_script)
    button.place(x=200, y=280)

def yelp_execute_script():
    try:
        td = entry1.get()
        td = td.replace(' ', '%20')
        name = "name={0}".format(td)
        address_text = entry2.get()
        address_text = address_text.split(',')
        address1 = address_text[0].split(' ')
        address = address1[:-2]
        address = " ".join(address)
        city = " ".join(address1[-2:])
        city = city.replace(' ', '%20')
        state = ''.join([i for i in address_text[1] if not i.isdigit()])
        state = state.strip()
        address = address.replace(' ', '%20')
        address = address.replace('#', "%23")
        country = 'US'
        yelp_url = "https://api.yelp.com/v3/businesses/matches?"
        address =  "&address1={0}".format(address)
        city  = "&city={0}".format(city)
        state  = "&state={0}".format(state)
        country = "&country={0}".format(country)
        yelp_api_key= "IEDZzra9jWr5BBJhZJkUV2_ck3ozOFjL8sIuCakAPWBqN6gKPN8R2wkliVO97WsCWp7QRos_P9m_ToRtTn4tpPaN77HkCyn1huepegGpVlqUNrAZCqwi5lKj8xw0Y3Yx"
        yelp_url = yelp_url + name + address + city + state + country +"&limit=3&match_threshold=default"
        print(yelp_url)
        yelp_headers = {'Authorization': 'Bearer %s' % yelp_api_key,}
        yelp_response = requests.request("GET",yelp_url, headers=yelp_headers, verify=False)
        place_id = yelp_response.json()['businesses'][0]['id']
        url = "https://api.yelp.com/v3/businesses/"
        part_url = place_id + "/reviews?limit=20&sort_by=yelp_sort"
        url = url + part_url
        response = requests.request('GET', url, headers=yelp_headers, verify=False)
        #response = requests.request('GET', url, headers=headers, params=url_params, verify=False)
        print(response.status_code)
        store_data = []
        response = response.json()['reviews']
        print(response)
        for i in response:
            processed_item = {}
            processed_item['Username'] = i['user']['name']
            processed_item['URL']  = i['url']
            processed_item['Comments']   = i['text']
            processed_item['Rating']  = i['rating']
            processed_item['Created Time'] = i['time_created']
            store_data.append(processed_item)
       
        df = pd.DataFrame(store_data)
        cols = list(df.columns)
        new_cols = []
        for col in cols:
            new_cols.append(col.capitalize())
        df.columns = new_cols
        output = entry3.get()
        if output.endswith('xlsx'):
            pass
        else:
            output =output +".xlsx"
        append_df_to_excel(output, df, 'Reviews', index = False)
        tkinter.messagebox.showinfo("Web scrapping result", "Successfully processed")  
         
    except:
        tkinter.messagebox.showinfo("Web scrapping result", "Reviews not found")
   
   

def yelp_reviews():
    global entry1, entry2, entry3
    label1 = Label(root,text="Enter company name",font=('arial',16),fg="black")
    label1.place(x=150, y=140)
    entry1 = Entry(root, width=40)
    entry1.place(x=150, y=180)
   
    label2 = Label(root,text="Enter the address",font=('arial',16),fg="black")
    label2.place(x=150,y=200)
    entry2 = Entry(root, width=40)
    entry2.place(x=150, y=240)
   
   
    label3 = Label(root,text="Enter name to save file",font=('arial',16),fg="black")
    label3.place(x=150, y=280)
    entry3 = Entry(root, width = 35)
    entry3.place(x=150, y=320)
   
    button = Button(root,width=10, text=" Download",command = yelp_execute_script)
    button.place(x=190, y=360)



def google_execute_script():
    given_url = entry2.get()
    print(given_url)
    if given_url.startswith('http') or given_url.startswith('www'):
        tsd, td, tsu = extract(given_url)
        td  = ' '.join(get_options(td, False, [], '')).capitalize()
    else:
        td = given_url
        print(td)
        td = td.replace(' ', "%20")
        url = "https://maps.googleapis.com/maps/api/place/findplacefromtext/json?input="
        url3= "&inputtype=textquery&fields=place_id&key=AIzaSyDMg-POYO8aPVIX4rOhUZqGus4GfdoMtXE"
        url = url+td+url3
        payload={}
        headers = {}
        response = requests.request("GET", url, headers=headers, data=payload)
        print(response.status_code)
        response = response.json()
        try:
            response = response['candidates'][0]['place_id']
            init_url= "https://maps.googleapis.com/maps/api/place/details/json?placeid="
            api_key = "&key=AIzaSyDMg-POYO8aPVIX4rOhUZqGus4GfdoMtXE"
            review_url = init_url + response + api_key
            review_response = requests.request("GET", review_url, headers=headers, data=payload)
            review_response = review_response.json()["result"]['reviews']
            df = pd.DataFrame(review_response)
            cols = list(df.columns)
            print('here')
            print(entry_3.get())
            new_cols = []
            for col in cols:
                new_cols.append(col.capitalize())
            df.columns = new_cols
            output = entry_3.get()
            if output.endswith('xlsx'):
                pass
            else:
                output =output +".xlsx"
            append_df_to_excel(output, df, 'Reviews', index = False)
            tkinter.messagebox.showinfo("Web scrapping result", "Successfully processed")  
               
        except:
            tkinter.messagebox.showinfo("Web scrapping result", "Reviews not found")

def google_reviews():
    label = Label(root,text="Please enter the Text",font=('arial',16),fg="black")
    label.place(x=150,y=140)
    global entry2, entry_3
    entry2 = Entry(root, width=40)
    entry2.place(x=150, y=190)
    label_2 = Label(root,text="Enter name for save file",font=('arial',16),fg="black")
    label_2.place(x=140,y=230)
    entry_3 = Entry(root, width = 35)
    entry_3.place(x=150, y=270)
    button = Button(root,width=10, text=" Download",command = google_execute_script)
    button.place(x=180, y=300)

def preprocessing():
    #print(text)
    #label.config(text = text)
    label1 = Label(root,text="Please enter the URL",font=('arial',16),fg="black", bg="lightgreen")
    label1.place(x=150,y=140)
    global entry1, entry2
    entry1 = Text(root, width=30, height=1)
    entry1.place(x=130, y=180)
    printButton = Button(root, width = 10,
                            text = "Search",
                            command = get_categories)
   
    printButton.place(x=380, y=175)

    label2 = Label(root, text= "Types of data available", font= ('arial', 16),fg="black", bg="lightgreen")
    label2.place(x=130,y=220)
    clicked = StringVar(root)
    label3 = Label(root,text="Enter name to save file",font=('arial',16),fg="black", bg="lightgreen")
    label3.place(x=130,y=300)
    entry2 = Entry(root, width = 35)
    entry2.place(x=130, y=340)
    downloadbutton = Button(root,width=10, text=" Download",command = execute_script)
    downloadbutton.place(x=180, y=370)

def show():
    text = clicked.get()
    removeable = []
    for ele in root.winfo_children():
        if len(removeable) >2:
            ele.destroy()
        else:
            removeable.append(ele.widgetName)
           
        #ele.widgetName), type(ele))
        #
    if text == "Web Scraping":
        preprocessing()
    if text == "Google reviews":
        google_reviews()
    if text == "Yelp reviews":
        yelp_reviews()
    if text == "Open Street Map":
        open_street_reviews()
       

def main():
    # Create object
    global root
    root = Tk()
    root.title("DataHarvest GPT")
    root.geometry("500x500")


    service = []
    cols = ["Services", "Description"]
    link_site = []
    h1 ='h1'
    h2 = 'h2'
    h3 ='h3'
    new_urls = []


    options = [
            "Web Scraping",
        "Yelp reviews",
        "Google reviews",
            "Open Street Map"
        ]

    # Set the background color to blue
    root.configure(bg='lightgreen')

    # Set the border color to red
    root.configure(bd=10, highlightbackground='red')
    txt = 'AI Web Data Parser'
    lbl = Label(root, text=txt, font='Bell 24 bold', fg='black', bg='orange', width=20)
    lbl.place(x=80,y=30)
    def animate_label(text, n=0):
        if n < len(text)-1:
            # not complete yet, schedule next run one second later
            lbl.after(1000, animate_label, text, n+1)
        # update the text of the label
        lbl['text'] = text[:n+1]

    # start the "after loop" one second later
    root.after(1000, animate_label, txt)
    c = options
    #root.configure( bg = 'white')
    global clicked
    clicked = StringVar()
    clicked.set('Select Service')
    main_menu = OptionMenu(root, clicked,  *c)
    #main_menu.config(font=helv36)
    main_menu.place(x= 150, y =85)
    font_properties = ('Arial', 10, 'bold')  # Font name, size, and style
    # Create a bold font for the default text
    bold_font = font.Font(family=font_properties[0], size=font_properties[1], weight='bold')
    main_menu.config(width = 12, font=bold_font)
    # Create button, it will change label text
    button = Button( root , text = "start" , width = 10,font='sans 10 bold', command = show )
    button.place(x=320, y=87)
    root.mainloop()
main()